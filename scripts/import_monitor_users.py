#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from google.cloud import firestore
from google.oauth2 import service_account

from app.contracts.admin import UserCreate
from app.domain.analysis_scopes import AnalysisScope, Department, SCOPE_POLICY_VERSION, membership_for
from app.domain.user_identity import roster_id_for_email
from app.repositories.user_directory import UserDirectoryRepository
from app.services.user_management import UserManagementService, area_key_for, normalize_roster_text
from app.settings import get_settings
try:
    from scripts.credential_preflight import approved_credential_path
except ModuleNotFoundError:
    from credential_preflight import approved_credential_path


REMARK_TO_DEPARTMENT = {
    "MR": Department.DM_FIELD,
    "本社(DM)": Department.DM_HQ,
    "本社(ヘルスケア)": Department.HEALTHCARE_HQ,
    "システム管理者": Department.ADMIN,
}
EXPECTED_HEADERS = (
    "エリア",
    "勤務地",
    "社員名",
    "社員メールアドレス",
    "役割",
    "テルモMR経歴",
    "備考",
)


@dataclass(frozen=True)
class RosterPlan:
    users: list[dict[str, Any]]
    scope_counts: dict[str, int]
    department_counts: dict[str, int]


def _canonical_remark(value: Any) -> str:
    return normalize_roster_text(str(value or "")).replace("（", "(").replace("）", ")")


def load_roster_plan(path: Path) -> RosterPlan:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    if "OurA-Naviユーザー管理" not in workbook.sheetnames:
        raise ValueError("required roster sheet is missing")
    sheet = workbook["OurA-Naviユーザー管理"]
    headers = [str(sheet.cell(1, index).value or "").strip() for index in range(1, 8)]
    if tuple(headers) != EXPECTED_HEADERS:
        raise ValueError(f"unexpected roster headers: {headers}")

    users: list[dict[str, Any]] = []
    emails: set[str] = set()
    for values in sheet.iter_rows(min_row=2, max_col=7, values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        row = dict(zip(headers, values))
        remark = _canonical_remark(row["備考"])
        try:
            department = REMARK_TO_DEPARTMENT[remark]
        except KeyError as exc:
            raise ValueError(f"unsupported roster remark: {remark}") from exc
        payload = UserCreate(
            name=normalize_roster_text(row["社員名"]),
            email=str(row["社員メールアドレス"] or ""),
            area=normalize_roster_text(row["エリア"]),
            workplace=normalize_roster_text(row["勤務地"]),
            role=normalize_roster_text(row["役割"]),
            department=department,
            mr_experience=(
                normalize_roster_text(row["テルモMR経歴"])
                if department is Department.DM_FIELD
                else "-"
            ),
            expected_scope_policy_version=SCOPE_POLICY_VERSION,
        )
        if payload.email in emails:
            raise ValueError("duplicate email in roster")
        emails.add(payload.email)
        users.append(
            {
                "roster_id": roster_id_for_email(payload.email),
                "user_id": "",
                "name": payload.name,
                "email": payload.email,
                "area": payload.area,
                "workplace": payload.workplace,
                "area_key": area_key_for(area=payload.area, workplace=payload.workplace),
                "role": payload.role,
                "department": department.value,
                "mr_experience": payload.mr_experience or "-",
                "label_ids": [],
                "chat_user_id": "",
                "is_active": True,
            }
        )

    memberships = [
        membership_for(
            role=item["role"],
            department=item["department"],
            is_active=item["is_active"],
        )
        for item in users
    ]
    return RosterPlan(
        users=users,
        scope_counts={
            AnalysisScope.GLOBAL.value: sum(item.includes(AnalysisScope.GLOBAL) for item in memberships),
            AnalysisScope.USER_MAP.value: sum(item.includes(AnalysisScope.USER_MAP) for item in memberships),
            "management": len(users),
        },
        department_counts=dict(Counter(item["department"] for item in users)),
    )


def _apply_plan(plan: RosterPlan, *, actor: str, credential_file: str) -> dict[str, Any]:
    settings = get_settings()
    credentials = service_account.Credentials.from_service_account_file(
        str(approved_credential_path(credential_file))
    )
    database = str(settings.monitor_firestore_database or "(default)").strip()
    client = firestore.Client(
        project=settings.monitor_project_id,
        database=database,
        credentials=credentials,
    )
    directory = UserDirectoryRepository(settings, client=client)
    existing = directory.list_users(include_inactive=True)
    planned_by_id = {str(item["roster_id"]): item for item in plan.users}
    existing_by_id = {
        str(item.get("roster_id") or ""): item
        for item in existing
    }
    unexpected_ids = sorted(set(existing_by_id) - set(planned_by_id))
    if unexpected_ids:
        raise RuntimeError(
            "roster import is bootstrap-only; unrelated production roster rows exist"
        )
    bootstrap_fields = (
        "roster_id", "user_id", "chat_user_id", "name", "email", "area",
        "workplace", "area_key", "role", "department", "mr_experience",
        "label_ids", "is_active",
    )
    for roster_id, current in existing_by_id.items():
        planned = planned_by_id[roster_id]
        mismatched = [
            field
            for field in bootstrap_fields
            if current.get(field) != planned.get(field)
        ]
        if mismatched:
            raise RuntimeError(
                "roster import is bootstrap-only; existing rows differ from the plan"
            )
    manager = UserManagementService(directory=directory)
    created = 0
    for item in plan.users:
        if item["roster_id"] in existing_by_id:
            continue
        manager.create_user(
            UserCreate(
                name=item["name"],
                email=item["email"],
                area=item["area"],
                workplace=item["workplace"],
                role=item["role"],
                department=item["department"],
                mr_experience=item["mr_experience"],
                expected_scope_policy_version=SCOPE_POLICY_VERSION,
            ),
            actor=actor,
        )
        created += 1
    applied = directory.list_users(include_inactive=True)
    expected_ids = {str(item["roster_id"]) for item in plan.users}
    applied_ids = {str(item.get("roster_id") or "") for item in applied}
    if applied_ids != expected_ids:
        raise RuntimeError("bootstrap readback does not match the planned roster")
    for current in applied:
        planned = planned_by_id[str(current.get("roster_id") or "")]
        if any(current.get(field) != planned.get(field) for field in bootstrap_fields):
            raise RuntimeError("bootstrap field readback does not match the planned roster")
    return {
        "appliedUsers": len(applied),
        "createdUsers": created,
        "resumedFromUsers": len(existing),
        "readbackMatched": True,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Plan or apply the canonical Monitor roster import")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--actor", default="")
    parser.add_argument("--credential-file")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()
    plan = load_roster_plan(args.workbook)
    summary = {
        "mode": "apply" if args.apply else "plan",
        "users": len(plan.users),
        "scopeCounts": plan.scope_counts,
        "departmentCounts": plan.department_counts,
    }
    print(json.dumps(summary, ensure_ascii=False, sort_keys=True))
    if not args.apply:
        return 0
    if not args.actor:
        raise SystemExit("--actor is required with --apply")
    receipt = _apply_plan(
        plan,
        actor=args.actor,
        credential_file=args.credential_file or "",
    )
    print(json.dumps({"applyReceipt": receipt}, ensure_ascii=False, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
